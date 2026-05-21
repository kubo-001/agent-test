# -*- coding: utf-8 -*-
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import sys
import os

def convert_md_to_excel(input_path, output_dir=None):
    """
    将Markdown测试用例转换为Excel格式

    Args:
        input_path: 输入的markdown文件路径
        output_dir: 输出目录，默认为 `output/` 目录
    """
    # 如果未指定输出目录，默认使用 output/
    if output_dir is None:
        # 从输入路径提取需求名称，构建 output/需求名称.xlsx 路径
        # 路径格式: .../{需求名称}_YYYY-MM-DD_HHmmss/test_cases_final.md
        input_dir = os.path.dirname(input_path)
        folder_name = os.path.basename(input_dir)  # 例如: 【阿虎医考App】v9.3.6 - POP付费资料_2026-05-14_100000

        # 去掉时间戳后缀，提取需求名称
        match = re.match(r'(.+)_\d{4}-\d{2}-\d{2}_\d{6}$', folder_name)
        if match:
            requirement_name = match.group(1)  # 例如: 【阿虎医考App】v9.3.6 - POP付费资料
        else:
            requirement_name = folder_name

        # Excel输出到 output/ 目录
        output_dir = 'output'
        # 确保output目录存在
        os.makedirs(output_dir, exist_ok=True)
        # Excel文件名 = 需求名称.xlsx
        excel_filename = f'{requirement_name}.xlsx'
        output_path = os.path.join(output_dir, excel_filename)
    else:
        # 如果指定了输出目录，使用目录+需求名称.xlsx
        folder_name = os.path.basename(os.path.dirname(input_path))
        match = re.match(r'(.+)_\d{4}-\d{2}-\d{2}_\d{6}$', folder_name)
        if match:
            requirement_name = match.group(1)
        else:
            requirement_name = folder_name
        excel_filename = f'{requirement_name}.xlsx'
        output_path = os.path.join(output_dir, excel_filename)

    print(f"读取文件: {input_path}")
    print(f"输出文件: {output_path}")

    # 读取文件
    with open(input_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # 去重：合并版本时可能产生重复的TC编号，只保留每个TC的最后一次出现
    lines = content.split('\n')
    seen_tcs = set()
    unique_lines = []
    for line in lines:
        case_match = re.match(r'^### (TC-\d+): (.+)', line)
        if case_match:
            tc_id = case_match.group(1)
            if tc_id in seen_tcs:
                continue  # 跳过旧版本
            seen_tcs.add(tc_id)
        unique_lines.append(line)
    content = '\n'.join(unique_lines)

    print(f"去重后保留 {len(seen_tcs)} 个唯一TC编号")

    # 解析测试用例
    test_cases = []
    current_feature = ""

    # 按行分割处理
    lines = content.split('\n')
    i = 0
    current_case = {}

    while i < len(lines):
        line = lines[i].strip()

        # 检查功能点标题
        feature_match = re.match(r'^## (F\d+): (.+)', line)
        if feature_match:
            current_feature = f"{feature_match.group(1)} {feature_match.group(2)}"
            i += 1
            continue

        # 检查测试用例标题
        case_match = re.match(r'^### (TC-\d+): (.+)', line)
        if case_match:
            if current_case.get('ID'):
                test_cases.append(current_case)
            case_id = case_match.group(1)
            case_title = case_match.group(2)
            current_case = {'ID': case_id, '标题': case_title, '目录': current_feature}
            i += 1
            continue

        # 检查字段 (需要去掉前导 - )
        clean_line = line[2:] if line.startswith('- ') else line  # 去掉 "- "
        if clean_line.startswith('**前置条件**:'):
            current_case['前置条件'] = clean_line.replace('**前置条件**:', '').strip()
        elif clean_line.startswith('**测试步骤**'):
            # 处理没有冒号的情况，如 "- **测试步骤**" 后直接跟步骤
            steps = ''
            # 合并多行步骤
            j = i + 1
            while j < len(lines):
                next_line = lines[j].strip()
                if not next_line:
                    break
                if next_line.startswith('**'):
                    break
                if next_line.startswith(('1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.', '10.')):
                    steps += '\n' + next_line
                    j += 1
                else:
                    break
            if steps:
                current_case['步骤描述'] = steps.strip()
            i = j
            continue
        elif clean_line.startswith('**测试步骤**:'):
            steps = clean_line.replace('**测试步骤**:', '').strip()
            # 合并多行步骤
            j = i + 1
            while j < len(lines):
                next_line = lines[j].strip()
                if not next_line:
                    break
                if next_line.startswith('**'):
                    break
                if next_line.startswith(('1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.', '10.')):
                    steps += '\n' + next_line
                    j += 1
                else:
                    break
            current_case['步骤描述'] = steps
            i = j
            continue
        elif clean_line.startswith('**步骤描述**:'):
            steps = clean_line.replace('**步骤描述**:', '').strip()
            # 合并多行步骤
            j = i + 1
            while j < len(lines):
                next_line = lines[j].strip()
                if not next_line:
                    break
                if next_line.startswith('**'):
                    break
                if next_line.startswith(('1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.', '10.')):
                    steps += '\n' + next_line
                    j += 1
                else:
                    break
            current_case['步骤描述'] = steps
            i = j
            continue
        elif clean_line.startswith('**预期结果**:'):
            current_case['预期结果'] = clean_line.replace('**预期结果**:', '').strip()
        elif clean_line.startswith('**优先级**:'):
            current_case['优先级'] = clean_line.replace('**优先级**:', '').strip()
        elif clean_line.startswith('**测试类型**:'):
            # 测试类型不写入Excel，跳过
            if current_case.get('标题'):
                test_cases.append(current_case)
            current_case = {}

        i += 1

    # 添加最后一个
    if current_case.get('标题'):
        test_cases.append(current_case)

    print(f"解析到 {len(test_cases)} 个测试用例")

    # 创建Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "testcase items"

    # 设置表头 - 按模板顺序
    headers = ['ID', '标题', '目录', '前置条件', '步骤描述', '预期结果', '负责人', '优先级']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_height = 20

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置表头行高
    ws.row_dimensions[1].height = header_height

    # 填充数据 - 按模板顺序
    for row, tc in enumerate(test_cases, 2):
        ws.cell(row=row, column=1, value=tc.get('ID', ''))
        ws.cell(row=row, column=2, value=tc.get('标题', ''))
        ws.cell(row=row, column=3, value=tc.get('目录', ''))
        ws.cell(row=row, column=4, value=tc.get('前置条件', ''))
        ws.cell(row=row, column=5, value=tc.get('步骤描述', ''))
        ws.cell(row=row, column=6, value=tc.get('预期结果', ''))
        ws.cell(row=row, column=7, value='')  # 负责人 - 留空
        ws.cell(row=row, column=8, value=tc.get('优先级', ''))

    # 设置列宽 - 根据模板调整
    ws.column_dimensions['A'].width = 10  # ID
    ws.column_dimensions['B'].width = 35  # 标题
    ws.column_dimensions['C'].width = 25  # 目录
    ws.column_dimensions['D'].width = 30  # 前置条件
    ws.column_dimensions['E'].width = 50  # 步骤描述
    ws.column_dimensions['F'].width = 50  # 预期结果
    ws.column_dimensions['G'].width = 12  # 负责人
    ws.column_dimensions['H'].width = 10  # 优先级

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 保存
    wb.save(output_path)
    print(f"已保存到: {output_path}")
    return output_path

if __name__ == '__main__':
    # 支持两种调用方式：
    # 1. python convert_to_excel.py <输入md路径>
    # 2. python convert_to_excel.py <输入md路径> <输出目录>
    if len(sys.argv) >= 2:
        input_path = sys.argv[1]
        output_dir = sys.argv[2] if len(sys.argv) >= 3 else None
        convert_md_to_excel(input_path, output_dir)
    else:
        print("用法: python convert_to_excel.py <输入md路径> [输出目录]")
        print("示例: python convert_to_excel.py ./test_cases_final.md")
        print("示例: python convert_to_excel.py ./test_cases_final.md ./output/")
